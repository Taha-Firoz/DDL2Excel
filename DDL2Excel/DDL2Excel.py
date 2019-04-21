import openpyxl
from pathlib import Path
from openpyxl.styles import Color, PatternFill, Alignment, Font, colors
import sys

from openpyxl import Workbook
if len(sys.argv) < 3 :
	print("Not enough arguments")
	exit()
if ".xlsx" not in sys.argv[2] and ".xls" not in sys.argv[2] :
	print("Invalid output file format")
	exit()
if ".txt" not in sys.argv[1]:
	print("Invalid input file format")
	exit()
config = Path(sys.argv[1])
if not config.is_file():
	print("Cannot find input file")
	exit()
	
wb = Workbook()
ws = wb.get_active_sheet()
current_row = 0
current_col = 1

table_heading = PatternFill(start_color='FF6AA84F',
                        end_color='FF6AA84F',
                        fill_type='solid')

table_columns = PatternFill(start_color='FFB6D7A8',
                         end_color='FFB6D7A8',
                         fill_type='solid')

f = open(sys.argv[1], "r")
is_table = False
for line in f:
    line = line.rstrip()

    if(");" in line):
        is_table = False
    if("create table" in line.lower()):
        is_table = True
        if(current_row != 0):
            cur_cell = ws.cell(row=current_row - 1, column=1)
            cur_cell.fill = table_heading
            cur_cell.font = Font(name="Ariel", color=colors.WHITE, bold=True)
            cur_cell.alignment = Alignment(
                horizontal='center', vertical='center')
            ws.merge_cells(start_row=current_row - 1, start_column=1,
                           end_row=current_row - 1, end_column=current_col - 1)
        current_row += 3
        current_col = 1
        line = line.replace("create", "").replace("Create", "").replace(
            "table", "").replace("Table", "").replace("(", "")
        ws.cell(row=current_row, column=1, value=line)
        current_row += 1

    elif(is_table and "constraint" not in line.lower() and ");" not in line and len(line) != 0):
        line = line.split()
        cur_cell = ws.cell(row=current_row, column=current_col)
        ws.cell(row=current_row, column=current_col, value=line[0])
        cur_cell.fill = table_columns
        cur_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_col += 1

if(current_row != 0):
    cur_cell = ws.cell(row=current_row - 1, column=1)
    cur_cell.fill = table_heading
    cur_cell.font = Font(name="Ariel", color=colors.WHITE, bold=True)
    cur_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=current_row - 1, start_column=1,
                   end_row=current_row - 1, end_column=current_col - 1)
f.close()
wb.save(sys.argv[2])
